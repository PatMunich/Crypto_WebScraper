import openpyxl as xl
from openpyxl.styles import PatternFill

class WorkbookEditor:
    def __init__(self):
        # Defines
        self.COLUMN_DATE = 1
        self.COLUMN_DATE_CHAR = 'A'
        self.COLUMN_WRITE_VALUE_SUBNETS = {'Targon': 'J', 'Nineteen': 'R', 'Gradients': 'Z', 'Chutes': 'AD'}
        self.DARK_GREEN = PatternFill(start_color='3cb371', end_color='3cb371', fill_type='solid')
        self.ERROR_CODES = {'1': "[WorkbookEditor error] during loading of the workbook!"}
        self.STATUS_CODES = {'1': f"[WorkbookEditor status] daily rewards successfully written into workbook!"}
        # Variables
        self.path_to_file = 'Files/Staking Rewards TAO.xlsx'
        try:
            self.workbook = xl.load_workbook(self.path_to_file)
        except:  # NOQA
            print(self.ERROR_CODES['1'])
        self.workbook_sheet_titles = self.getWorksheetTitles()
        self.timestamp_day_month = ''

    def getWorksheetTitles(self):  # NOQA
        workbook_sheets = {}
        for idx in range(len(self.workbook.worksheets)):
            workbook_sheets[idx] = self.workbook.worksheets[idx].title
        return workbook_sheets

    def sortInSubnetData(self, in_subnet_data):  # NOQA
        for idx in self.workbook_sheet_titles:
            if in_subnet_data['0'] in self.workbook_sheet_titles[idx]:
                target_sheet = self.workbook.get_sheet_by_name(in_subnet_data['0'])  # NOQA
                for row_id in range(target_sheet.max_row):
                    cell_content = target_sheet.cell(row=row_id + 1, column=self.COLUMN_DATE).value
                    try:
                        cell_content = cell_content.strftime('%Y-%d/%m')
                        timestamp_day_month = cell_content.split('-')[1]
                        if timestamp_day_month == in_subnet_data['1']:
                            in_subnet_data.pop('0')
                            in_subnet_data.pop('1')
                            for subnet in in_subnet_data:
                                write_coordinate = str(self.COLUMN_WRITE_VALUE_SUBNETS[subnet]) + str(row_id + 1)
                                subnet_balance = in_subnet_data[str(subnet)]
                                try:
                                    subnet_balance = in_subnet_data[str(subnet)].replace('.', ',')
                                except:  # NOQA
                                    pass
                                target_sheet[write_coordinate] = subnet_balance
                                target_sheet[self.COLUMN_DATE_CHAR + str(row_id + 1)].fill = self.DARK_GREEN
                            self.workbook.save(self.path_to_file)
                            print(self.STATUS_CODES['1'])
                            return
                    except:  # NOQA
                        pass
